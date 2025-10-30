# -*- coding: utf-8 -*-
import re
import time
import os
import requests
import pandas as pd
import xlwings as xw
import openpyxl as px

from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from urllib.parse import urlparse, parse_qs, unquote
from io import StringIO
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from datetime import datetime


# ------------ レース情報抽出（共通）------------
def race_info(url):
    headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
                  }
    race_infos = []
    r = requests.get(url, headers=headers, timeout=20)
    # 文字化け対策
    if r.apparent_encoding:
        r.encoding = r.apparent_encoding

    soup = BeautifulSoup(r.text, "lxml")

    # 1) レース名（例: 西湖特別）
    race_name = soup.select_one(".RaceName")
    race_name = race_name.get_text(strip=True) if race_name else None

    # 2) 開催時刻/コース/天候/馬場（例: 16:50発走 / ダ1600m (左) / 天候:曇 / 馬場:良）
    race_data01 = soup.select_one(".RaceData01")
    race_data01 = re.sub(r"\s+", " ", race_data01.get_text(" ", strip=True)) if race_data01 else None

    # 3) 開催情報 + クラス等（例: 4回 東京 7日目 サラ系３歳以上 ２勝クラス …）
    race_data02 = soup.select_one(".RaceData02")
    race_data02 = re.sub(r"\s+", " ", race_data02.get_text(" ", strip=True)) if race_data02 else None
    
    race_infos = []
    if not race_name :
        raise ValueError("race_infosにrace_nameが存在しません。")
    elif not race_data01 :
        raise ValueError("race_infosにrace_data01が存在しません。")
    elif not race_data02 :
        raise ValueError("race_infosにrace_data02が存在しません。")
    else:
        race_infos = [race_name,race_data01,race_data02]
    return race_infos


# ------------ 出馬表抽出（共通）------------
def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    if isinstance(df.columns, pd.MultiIndex):
        flat = []
        for tup in df.columns:
            parts = [str(x).strip() for x in tup if str(x).strip()]
            s = " ".join(parts) if parts else ""
            toks = s.split()
            if len(toks) >= 2 and len(set(toks)) == 1:
                s = toks[0]
            flat.append(s)
        df.columns = flat
    else:
        df.columns = [str(c).strip() for c in df.columns]

    # 重複列名に連番
    seen = {}
    uniq = []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            uniq.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            uniq.append(c)
    df.columns = uniq

    # 先頭行がヘッダ複写なら落とす
    if len(df) > 0:
        row0 = df.iloc[0].astype(str).str.replace(r"\s+", "", regex=True).tolist()
        col0 = [c.replace(" ", "") for c in df.columns]
        match_cnt = sum(1 for x in row0 if any(x and x in c for c in col0))
        if match_cnt >= max(2, len(col0)//2):
            df = df.iloc[1:].reset_index(drop=True)
    return df

def _decode_html(bytes_data: bytes, fallback: str = "utf-8") -> str:
    # """
    # HTMLバイト列 -> 正しいUnicode文字列へ。
    # metaのcharsetや自動判定を使って確実にデコードする。
    # """
    # 先頭4KBだけ見て <meta charset=...> を拾う
    head = bytes_data[:4096]
    m = re.search(rb"charset\s*=\s*['\"]?\s*([A-Za-z0-9_\-]+)", head, re.I)
    enc = m.group(1).decode("ascii", "ignore").lower() if m else None

    # EUC-JP の別名吸収
    if enc in {"eucjp", "euc_jp", "ujis"}:
        enc = "euc-jp"

    # 見つからなければ自動判定（charset-normalizer / chardet）
    if not enc:
        try:
            from charset_normalizer import from_bytes
            enc = from_bytes(bytes_data).best().encoding
        except Exception:
            enc = None

    if not enc:
        enc = fallback

    # 最後にデコード
    return bytes_data.decode(enc, errors="replace")


# def _ensure_unicode(html_or_bytes) -> str:
#     """
#     文字化け救済: bytes なら正規にデコード。
#     str で“化けて”いる可能性が高いときは、latin-1 でバイトに戻して再デコードを試す。
#     """
#     if isinstance(html_or_bytes, bytes):
#         return _decode_html(html_or_bytes, fallback="utf-8")

#     # すでに str の場合
#     text: str = html_or_bytes
#     # “��”が多い・EUCページでありがちな記号化けが多い → 化けてると判断
#     if ("��" in text) or ("EUC-JP" in text and re.search(r"[^\x00-\x7F]", text) is False):
#         # いったんバイトに戻す（元の生バイトを“そのまま”取り出すには latin-1 が有効）
#         try:
#             raw = text.encode("latin-1", errors="ignore")
#             # meta から再デコード（EUC-JPが多い）
#             return _decode_html(raw, fallback="euc-jp")
#         except Exception:
#             pass
#     return text

def _extract_from_html(html: str) -> pd.DataFrame | None:
    html_io = StringIO(html)
    try:
        tables = pd.read_html(html_io, flavor="lxml")
    except Exception:
        html_io.seek(0)
        tables = pd.read_html(html_io)
        
    # ★「オッズ」を必須にしないように変更（5列は必須、オッズは任意）
    REQUIRED = {"馬番", "人気順", "馬名", "騎手名", "斤量"}  # ※人気順は後で「人気」にもマッチさせる
    OPTIONAL = {"オッズ"}
    
    col_patterns = {
        "馬番":   re.compile(r"(馬\s*番|枠\s*番|馬番|枠番|\b馬\s*#?)", re.I),
        "人気順": re.compile(r"(人気|単勝人気)", re.I),
        "オッズ": re.compile(r"(オッズ|単勝)", re.I),
        "馬名":   re.compile(r"(馬\s*名|馬名|名前)", re.I),
        "騎手名": re.compile(r"(騎手|騎手名|ジョッキー)", re.I),
        "斤量":   re.compile(r"(斤量|負担重量|負担重|重量)", re.I),
    }

    def pick(df: pd.DataFrame) -> pd.DataFrame | None:
        df = _normalize_columns(df)
        cols = [str(c) for c in df.columns]
        mapping = {}
        for want, pat in col_patterns.items():
            hit = next((c for c in cols if pat.search(c)), None)
            if hit:
                mapping[hit] = want

        # 代替マッピング（足りないときだけ補完）
        if len(set(mapping.values())) < 6:
            for c in cols:
                if re.search(r"(印|予想印)", c) and "人気順" not in mapping.values():
                    mapping[c] = "人気順"
                if re.search(r"(単勝|勝率)", c) and "オッズ" not in mapping.values():
                    mapping[c] = "オッズ"
                if re.search(r"(騎手|ジョッキー)", c) and "騎手名" not in mapping.values():
                    mapping[c] = "騎手名"
                if re.search(r"(斤量|負担重量|負担重|重量)", c) and "斤量" not in mapping.values():
                    mapping[c] = "斤量"

        if len(set(mapping.values())) == 6:
            out = df[list(mapping.keys())].rename(columns=mapping).copy()

            out["人気順"] = pd.to_numeric(
                out["人気順"].astype(str).str.extract(r"(\d+)", expand=False),
                errors="coerce"
            )

            out["オッズ"] = (
                out["オッズ"].astype(str)
                .str.replace("倍", "", regex=False)
                .str.replace(",", "", regex=False)
            )
            out["オッズ"] = pd.to_numeric(out["オッズ"], errors="coerce")

            out["馬番"] = pd.to_numeric(
                out["馬番"].astype(str).str.extract(r"(\d+)", expand=False),
                errors="coerce"
            ).astype("Int64")

            out["騎手名"] = (
                out["騎手名"].astype(str)
                .str.replace(r"\s+", " ", regex=True)
                .str.strip()
            )

            out["斤量"] = pd.to_numeric(
                out["斤量"].astype(str).str.extract(r"(\d+(?:\.\d+)?)", expand=False),
                errors="coerce"
            )

            out["馬名"] = out["馬名"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()

            # 並べ替え（馬番→人気順にしたければ ["馬番","人気順"] に）
            out = out.sort_values(["人気順", "馬番"], na_position="last", ignore_index=True)
            out = out[out["馬名"].str.len() > 0].reset_index(drop=True)
            return out
        return None

    for tb in tables:
        got = pick(tb)
        if got is not None:
            return got
    return None



# ------------ ここが重要：Seleniumのタイムアウト対策 ------------
def _new_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service as ChromeService
    from webdriver_manager.chrome import ChromeDriverManager
    import os
    os.environ["WDM_LOG"] = "0"        # webdriver_manager のログ抑制
    os.environ["WDM_PRINT_FIRST_LINE"] = "False"
    options = webdriver.ChromeOptions()
    # 既存の軽量化オプション
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-software-rasterizer")
    options.add_argument("--disable-webgl")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--no-sandbox")
    options.add_argument("--blink-settings=imagesEnabled=false")
    options.add_argument("--lang=ja-JP")
    options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                         "KHTML, like Gecko) Chrome/120.0 Safari/537.36")
    options.page_load_strategy = "none"

    # ★ これが効く：ws出力を止める（portではなくpipeでDevTools接続）
    options.add_argument("--remote-debugging-pipe")
    
    # （ログ抑制の保険）
    options.add_experimental_option("excludeSwitches", ["enable-logging"])
    options.add_argument("--log-level=3")
    options.add_argument("--silent")
    os.environ["WDM_LOG"] = "0"
    os.environ["WDM_PRINT_FIRST_LINE"] = "False"
    os.environ["CHROME_LOG_FILE"] = os.devnull  # Chrome独自ログの行き先も捨てる

    # ★ 追加：ChromeDriver 側のログも捨てる
    service = ChromeService(
        ChromeDriverManager().install(),
        log_output=open(os.devnull, "w", encoding="utf-8", errors="ignore")
    )

    driver = webdriver.Chrome(service=service, options=options)
    driver.set_page_load_timeout(60)
    driver.set_script_timeout(60)
    return driver


def _get_rendered_html(url: str, wait_css: str = None, hard_timeout: int = 20, wait_odds: bool = False) -> str:
    driver = _new_driver()
    try:
        try:
            driver.get(url)
        except TimeoutException:
            pass

        # DOMContentLoaded 相当まで待機
        t0 = time.time()
        while time.time() - t0 < hard_timeout:
            ready = driver.execute_script("return document.readyState")
            if ready in ("interactive", "complete"):
                break
            time.sleep(0.2)

        # ★ 出馬表タブがあればクリック（「出馬表」「出馬」「枠順」などを許容）
        try:
            # クリック候補(テキスト)を広めに
            candidates = ["出馬表", "出馬", "枠順"]
            tabs = driver.find_elements(By.CSS_SELECTOR, "a, button")
            for el in tabs:
                try:
                    txt = (el.text or "").strip()
                    if any(k in txt for k in candidates):
                        el.click()
                        time.sleep(0.3)
                        break
                except Exception:
                    pass
        except Exception:
            pass
        
        # まずはテーブル本体や行の出現まで待つ
        if wait_css:
            WebDriverWait(driver, hard_timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, wait_css))
            )
        else:
            # 代表的なテーブル/行
            selectors = [
                ".Shutuba_Table", "table.Shutuba_Table",
                ".RaceTable01", "table.RaceTable01",
                ".Shutuba_Table tbody tr", ".RaceTable01 tbody tr",
            ]
            end = time.time() + hard_timeout
            ok = False
            while time.time() < end and not ok:
                for sel in selectors:
                    try:
                        WebDriverWait(driver, 1).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                        )
                        ok = True
                        break
                    except Exception:
                        pass
                if not ok:
                    driver.execute_script("window.scrollBy(0, 400);")
                    time.sleep(0.2)

        # ★ ここが重要：オッズの“数値”が入るまで待つ（任意）
        if wait_odds:
            def odds_ready(drv):
                try:
                    return drv.execute_script("""
                        const nodes = document.querySelectorAll('td.Popular, td.Odds, .Popular, .Odds');
                        for (const n of nodes) {
                          const t = (n.textContent || '').trim();
                          if (/^\\d+(?:\\.\\d+)?$/.test(t)) return true; // 例: 1.7 / 12 / 3.9
                          // "1.7倍" のような表記も許可
                          if (/\\d+(?:\\.\\d+)?\\s*倍/.test(t)) return true;
                        }
                        return false;
                    """)
                except Exception:
                    return False

            WebDriverWait(driver, hard_timeout).until(lambda d: odds_ready(d))

        # まだ完了でなくても現状を確保
        if driver.execute_script("return document.readyState") != "complete":
            driver.execute_script("window.stop();")

        return driver.page_source
    finally:
        driver.quit()


def fetch_shutsuba(url: str, use_selenium_fallback: bool = True, timeout_sec: int = 20) -> pd.DataFrame:
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": "https://race.netkeiba.com/",
        "Accept-Language": "ja,en;q=0.8",
    }
    r = requests.get(url, headers=headers, timeout=timeout_sec)
    r.raise_for_status()

    # ★ ここがポイント：バイト列から確実にデコード
    html_text = _decode_html(r.content, fallback="utf-8")
    df = _extract_from_html(html_text)
    if df is not None and len(df) > 0:
        return df

    if not use_selenium_fallback:
        raise ValueError("出馬表テーブルが見つかりません（静的HTML）。Seleniumフォールバックを有効にしてください。")

    # Selenium 側はブラウザが正しくデコードしてくれるので、そのまま Unicode 文字列
    html_after = _get_rendered_html(
        url,
        wait_css=".Shutuba_Table, table.RaceTable01, .RaceTable01",   # ← 追加
        hard_timeout=60, # ← 余裕を持たせる
        wait_odds=True
    )
    df2 = _extract_from_html(html_after)
    if df2 is not None and len(df2) > 0:
        return df2

    # ---- ダンプして原因を見える化（例外前に保存）----
    _dump_debug_html(html_after, url)   # ← 追加（下に関数あり）

    # 少しスクロール→再取得のところも同様
    driver = _new_driver()
    try:
        try:
            driver.get(url)
        except TimeoutException:
            pass
        for y in (400, 800, 1200, 1600, 2000, 2800):
            driver.execute_script(f"window.scrollTo(0,{y});")
            time.sleep(0.4)
        driver.execute_script("window.stop();")
        html_after = driver.page_source  # ここは既に正しい文字列
    finally:
        driver.quit()

    df3 = _extract_from_html(html_after or "")
    if df3 is not None and len(df3) > 0:
        return df3
    
    _dump_debug_html(html_after, url)   # ← 最後にも保存
    raise ValueError("出馬表テーブルが見つかりません（Seleniumでも取得不可）。")


def sort_shutsuba(df, by="馬番"):
    if by == "馬番":
        return df.sort_values(["馬番", "人気順"], na_position="last", ignore_index=True)
    elif by == "人気順":
        return df.sort_values(["人気順", "馬番"], na_position="last", ignore_index=True)
    else:
        return df
    
def output_fileName(url):
    # race_id= の値を取得
    race_infos = race_info(url)
    race_name = race_infos[0]
    # q = parse_qs(urlparse(url).query)
    # rid = unquote(q.get('race_id', [''])[0])  # 例: '202508030711'

    # 出力ファイル名に使う（必要なら拡張子を変える）
    outfile_csv  = f"{race_name}.csv"
    # outfile_xlsx = f"shutsuba_{rid}.xlsx"

    # 任意の保存先（例）
    save_dir = r"D:\python\program\data"
    path_csv  = os.path.join(save_dir, outfile_csv)
    # path_xlsx = os.path.join(save_dir, outfile_xlsx)
    return path_csv

# # ---- ヘッダーから「場所」「R」「距離」「芝/ダ」「左/右」を抜いてシート名を作る ----
# def _sheet_name_from_html(html: str, fallback_race_id: str = "") -> str:
#     """
#     ページの本文テキストから
#       例: 「東京10R ダ1600m (左)」/「京都10R 芝1400m（右）」 等を検出して
#     シート名「東京10R,1600メートル（ダート・左）」の形で返す。
#     取れない場合は「race_{race_id}」を返す。
#     """
    
#     # ★ここがポイント：化け対策
#     text = _ensure_unicode(html)
#     text = re.sub(r"\s+", " ", text)  # 改行や連続空白を1スペースに
    
#     # コース名 + レース番号
#     m1 = re.search(r"([一-龥]{1,3})\s*?(\d{1,2})R", text)  # 例: 東京10R / 京都10R / 中山11R など
#     # 芝/ダ + 距離m + (左|右)
#     m2 = re.search(r"(芝|ダ)\s*?(\d{3,4})\s*[mｍ]\s*?[（(]\s*(左|右)\s*[）)]", text)

#     if not m1 or not m2:
#         return f"race_{fallback_race_id}" if fallback_race_id else "race_unknown"

#     place = m1.group(1)            # 東京/京都/中山/阪神/新潟/札幌/函館/中京/小倉 等
#     race_no = int(m1.group(2))     # 10 など
#     sd = m2.group(1)               # 芝 or ダ
#     dist = int(m2.group(2))        # 1600 など
#     lr = m2.group(3)               # 左 or 右

#     surface = "芝" if sd == "芝" else "ダート"
#     # 完成形: 「東京10R,1600メートル（ダート・左）」
#     return f"{place}{race_no}R,{dist}メートル（{surface}・{lr}）"

def safe_sheet_name(name: str, used: set[str]) -> str:
    # """
    # Excelの制約対応 & 重複回避
    #   - 禁止文字 \ / * ? : [ ]
    #   - 長さ31文字まで
    #   - 既に使われている場合は _2, _3 ... を付与
    # """
    base = re.sub(r"[\\/*?:\[\]]", "_", name).strip()
    if not base:
        base = "sheet"
    base = base[:31]

    candidate = base
    i = 2
    # 重複していたら末尾に連番を付ける。31文字を超えるなら末尾を削って調整。
    while candidate in used:
        suffix = f"_{i}"
        candidate = (base[: max(0, 31 - len(suffix))] + suffix)[:31]
        i += 1
    used.add(candidate)
    return candidate

#エクセルに罫線を加える
def all_sheets_orgformat(filename,headercolorflg):
    #黒の細線で罫線を引く
    side1 = Side(style='thin',color = '000000')
    border1 = Border(top=side1,bottom=side1,left=side1,right=side1)
    wb = px.load_workbook(filename)
    #すべてのシート
    for ws in wb.worksheets:
        #データが入っているすべてのセル
        for row in ws.rows:
            for cell in row:
                #ヘッダがある場合はTrueにする
                if headercolorflg:
                    if cell.row == 1:
                        #ヘッダーは、黄色にする
                        cell.fill = PatternFill(fgColor='FFFF00',bgColor='FFFF00',fill_type='solid')
                    if cell.row %2 == 0:
                        #ヘッダ以外で偶数行の場合は,薄い黄色にする
                        cell.fill = PatternFill(fgColor='FFFFE0',bgColor='FFFFE0',fill_type='solid')
                    cell.border = border1

    wb.save(filename)

# 失敗時にHTMLを保存して原因を確認
def _dump_debug_html(html: str, url: str):
    try:
        if not html:
            return
        os.makedirs(r"D:\python\program\data\debug", exist_ok=True)
        rid = parse_qs(urlparse(url).query).get("race_id", ["unknown"])[0]
        path = rf"D:\python\program\data\debug\debug_{rid}.html"
        with open(path, "w", encoding="utf-8", errors="ignore") as f:
            f.write(html)
        print(f"[DEBUG] saved HTML: {path}")
    except Exception as _:
        pass


#実行
if __name__ == "__main__":
    errors = []
    written = 0   
    
    #抽出したいレースのレースIDを以下リストに入れる
    race_ids = [202505040910,202508030910,202504040411,202505040911,202508030911]
    #出力ファイルの末尾につく日付けを入れる
    race_date = 20251026
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    # 1つのExcelにまとめて出力するパスを決める
    save_dir = r"D:\python\program\data"
    os.makedirs(save_dir, exist_ok=True)
    output_xlsx = os.path.join(save_dir, f"Win5レース({race_date})_{now}.xlsx")
    print(f"出力開始: {output_xlsx}")
    
    used_sheet_names: set[str] = set()
    num_race = 1
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        for race_id in race_ids:
            url = f"https://race.netkeiba.com/race/shutuba.html?race_id={race_id}"
            try:
            
                race_infos = race_info(url)
                race_name = race_infos[0]
                sheet_name = safe_sheet_name(race_name, used_sheet_names)

                print(f"第{written + 1}レース [{sheet_name}]シートに書き込み中…")

                # ★ テーブル待機CSSを明示してから抽出
                df = fetch_shutsuba(url)
                df = sort_shutsuba(df, by="馬番")

                df.to_excel(writer, index=False, sheet_name=sheet_name)
                print(f"第{written + 1}レース [{sheet_name}]シートに書き込み完了")
                written += 1

            except Exception as e:
                msg = f"{race_id}: {type(e).__name__}: {e}"
                print("[SKIP]", msg)
                errors.append(msg)
                continue

        # 1枚も書けなかったときは空シートを作って Excel を壊さない
        if written == 0:
            pd.DataFrame({"info": ["no sheets written"]}).to_excel(
                writer, index=False, sheet_name="empty"
            )

        # 失敗があったらログシートも残す（任意）
        if errors:
            pd.DataFrame({"errors": errors}).to_excel(
                writer, index=False, sheet_name="log"
            )
    all_sheets_orgformat(output_xlsx,True)
    print(f"出力完了: {output_xlsx}")