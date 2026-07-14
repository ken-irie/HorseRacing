# -*- coding: utf-8 -*-
import os
import re
import sys
import time
import datetime as dt
import requests

from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from bs4 import BeautifulSoup
from bs4 import UnicodeDammit
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ===================== 定数 =====================
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36"),
    "Referer": "https://www.netkeiba.com/",
    "Accept-Language": "ja,en;q=0.9",
}
# 土曜日はidx=0、日曜日はidx=1（launcher.pyからは環境変数WIN5_IDXで上書きされる）
try:
    idx = int(os.environ.get("WIN5_IDX", "1"))
except ValueError:
    idx = 1
PC_URL = f"https://race.netkeiba.com/top/win5.html?idx={idx}"
RACE_ID_RE = re.compile(r"race_id=(\d{12})")

# テンプレートファイル（スクリプトと同じフォルダに置く）
TEMPLATE_XLSX = Path(__file__).resolve().with_name("main_horse_decide_sheets.xlsx")
# ===================== 定数 =====================

# ===================== 高速化：HTTPセッション =====================
def build_session() -> requests.Session:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    s = requests.Session()
    s.headers.update(HEADERS)
    retry = Retry(
        total=3, backoff_factor=0.3,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",)
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://", HTTPAdapter(max_retries=retry))
    return s

SESSION = build_session()
# ===================== 高速化：HTTPセッション =====================

# ===================== HTMLユーティリティ =====================
def _decode_html_bytes(b: bytes, fallback: str = "utf-8") -> str:
    dammit = UnicodeDammit(b, is_html=True)
    if dammit.unicode_markup:
        return dammit.unicode_markup
    return b.decode(fallback, errors="replace")

def _get_html(url: str, timeout: int = 15) -> str:
    r = SESSION.get(url, timeout=timeout)
    r.raise_for_status()
    return _decode_html_bytes(r.content)
# ===================== HTMLユーティリティ =====================

# ===================== WIN5 race_idとrace_date 抽出 =====================
def _extract_ids_from_html(html: str) -> list[str]:
    soup = BeautifulSoup(html, "lxml")
    ids, seen = [], set()
    for a in soup.find_all("a", href=True):
        m = RACE_ID_RE.search(a["href"])
        if not m:
            continue
        rid = m.group(1)
        if rid not in seen:
            seen.add(rid)
            ids.append(rid)
    return ids

def _race_date(html: str) -> str:
    soup = BeautifulSoup(html, "lxml")

    # 年は WIN5ページ中の race_id=YYYY…… から取得
    year = ""
    ids = _extract_ids_from_html(html)
    if ids:
        year = ids[0][:4]

    active_dd = soup.select_one(".RaceList_Date dl.Win5_Date dd.Active")
    if not active_dd:
        active_dd = soup.select_one(".RaceList_Date dd.Active")
    if not active_dd:
        return ""

    txt = active_dd.get_text(" ", strip=True)
    m_md = re.search(r"(\d{1,2})月\s*(\d{1,2})日", txt)
    if not m_md:
        return ""

    month = int(m_md.group(1))
    day = int(m_md.group(2))

    if not year:
        page_text = soup.get_text(" ", strip=True)
        m_y = re.search(r"(\d{4})年", page_text)
        if m_y:
            year = m_y.group(1)
        else:
            year = str(dt.datetime.now().year)

    date = f"{int(year):04d}{month:02d}{day:02d}"
    return date

def pick_win5_ids(target_url: str | None = None):
    url = target_url or PC_URL
    try:
        html = _get_html(url)
        ids = _extract_ids_from_html(html)
        date = _race_date(html)
    except Exception:
        return [], ""

    if len(ids) >= 5:
        return ids[:5], date
    return ids, date
# ===================== WIN5 race_idとrace_date 抽出 =====================

def parse_past_cell(td) -> tuple[str, str, str, str, str, str, str]:
    """
    過去走1つ分の <td class="Past"> から
    (レース名, 場所, コース, 着順, 着差, 通過順, ３F) を取り出す
    """
    if td is None:
        return "", "", "", "", "", "", ""

    # レース名（aタグ直下テキストのみ）
    race_name = ""
    a_tag = td.select_one("div.Data02 a")
    if a_tag:
        text = a_tag.find(string=True, recursive=False)
        if text:
            race_name = text.strip()

    # 日付＋場所 → 場所だけ抜く
    place = ""
    span_day_place = td.select_one("div.Data01 span:not(.Num)")
    if span_day_place:
        t = span_day_place.get_text(strip=True)
        # 例: "2025.09.15 阪神" → 最後の要素を場所とみなす
        parts = t.split()
        if len(parts) >= 2:
            place = parts[-1]

    # コース（芝1600, 芝1600(外) など）
    course = ""
    div_course = td.select_one("div.Data05")
    if div_course:
        t = div_course.get_text(" ", strip=True)
        # "芝1600 1:36.2 良" のような文字列 → 最初の要素だけ
        parts = t.split()
        if parts:
            course = parts[0]

    # 着順（Data01内の span.Num）
    finish = ""
    span_num = td.select_one("div.Data01 span.Num")
    if span_num:
        finish = span_num.get_text(strip=True)

    # 着差（Data07 内の (...)）
    margin = ""
    div_margin = td.select_one("div.Data07")
    if div_margin:
        t = div_margin.get_text(" ", strip=True)
        m = re.search(r"\(([^)]+)\)", t)   # 例: "(0.3)" → "0.3"
        if m:
            margin = m.group(1).strip()

    # Data06（通過順 + 3Fが入ってる想定）
    passing = ""
    last3f = ""
    div_06 = td.select_one("div.Data06")
    if div_06:
        t = div_06.get_text(" ", strip=True)  # 例: "4-3-4-3 (33.9) 524(+10)"

        # 通過順：先頭の "4-3-4-3" を取る（括弧の前）
        # パターンが崩れても split の先頭で拾えるようにする
        if t:
            first = t.split()[0]
            # 先頭が "4-3-4-3" 形式のときだけ採用（安全策）
            if re.fullmatch(r"\d+(?:-\d+)+", first):
                passing = first

        # 3F：括弧内数値
        m = re.search(r"\(([\d\.]+)\)", t)
        if m:
            last3f = m.group(1)

    return race_name, place, course, finish, margin, passing, last3f

# ===================== レースメタ情報抽出 =====================
def _extract_race_meta(html: str) -> tuple[str, str, str, str, str, str]:
    soup = BeautifulSoup(html, "lxml")

    name_el = soup.select_one(".RaceName")
    name = name_el.get_text(strip=True) if name_el else ""

    d1_el = soup.select_one(".RaceData01")
    d2_el = soup.select_one(".RaceData02")
    d1 = d1_el.get_text(" ", strip=True) if d1_el else ""
    d2 = d2_el.get_text(" ", strip=True) if d2_el else ""

    year = ""
    m_id = RACE_ID_RE.search(html)
    if m_id:
        rid = m_id.group(1)
        year = rid[:4]

    race_date = ""
    active_dd = soup.select_one("#RaceList_DateList dd.Active")
    if active_dd and year:
        txt = active_dd.get_text(" ", strip=True)
        m_md = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", txt)
        if m_md:
            month = int(m_md.group(1))
            day = int(m_md.group(2))
            race_date = f"{int(year):04d}{month:02d}{day:02d}"

    if not race_date:
        date_text_candidates: list[str] = []
        date_el = soup.select_one(".RaceList_Date")
        if date_el:
            date_text_candidates.append(date_el.get_text(" ", strip=True))
        if d1:
            date_text_candidates.append(d1)
        if d2:
            date_text_candidates.append(d2)

        for txt in date_text_candidates:
            m = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", txt)
            if m:
                y, mth, d = m.groups()
                race_date = f"{int(y):04d}{int(mth):02d}{int(d):02d}"
                break

    place = ""
    if d2:
        PLACE_PATTERN = re.compile(r"(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)")
        m = PLACE_PATTERN.search(d2)
        if m:
            place = m.group(1)

    rnum = ""
    rnum_el = soup.select_one(".RaceNum")
    rnum_text = rnum_el.get_text(strip=True) if rnum_el else ""
    m_r = re.search(r"(\d+)R", rnum_text)
    if not m_r:
        m_r = re.search(r"(\d+)R", soup.get_text(" ", strip=True))
    if m_r:
        rnum = m_r.group(1) + "R"

    return race_date, name, place, rnum, d1, d2


def _parse_race_time(d1: str) -> str:
    """RaceData01から発走時刻（HH:MM）を抽出する"""
    m = re.search(r"(\d{1,2}:\d{2})", d1 or "")
    return m.group(1) if m else ""


def _parse_course_label(d1: str, d2: str) -> str:
    """〇歳以上〇勝クラス〇メートル（芝・右）形式の文字列を組み立てる"""
    d1 = d1 or ""
    d2 = d2 or ""

    # 距離（例: 1,800m / 3000m）
    m = re.search(r"([\d,]+)\s*m", d1)
    dist = f"{m.group(1)}メートル" if m else ""

    # 芝/ダート と 右/左/直線（netkeiba表記は「ダ1700m (右)」「芝1800m (右 A)」など）
    m_surf = re.search(r"(障害|障|ダート|ダ|芝)", d1)
    m_dir  = re.search(r"[(（]\s*(右|左|直線)", d1)
    surf_map = {"芝": "芝", "ダ": "ダート", "ダート": "ダート", "障": "障害", "障害": "障害"}
    surf = surf_map.get(m_surf.group(1), "") if m_surf else ""
    drct = m_dir.group(1) if m_dir else ""
    if surf and drct:
        course = f"（{surf}・{drct}）"
    elif surf:
        course = f"（{surf}）"
    else:
        course = ""

    # 年齢条件（例: 4歳以上 / 3歳）とクラス条件をd2から抽出
    m_age = re.search(r"(\d+歳[以上未満]*)", d2)
    age = m_age.group(1) if m_age else ""
    m_cls = re.search(r"(\d+勝クラス|オープン|G[IⅠ1iIVX]{1,3}|ハンデ|混合|牝馬限定)", d2)
    cls = m_cls.group(1) if m_cls else ""

    return f"{age}{cls}{dist}{course}".strip()

# ===================== サイトからデータ取得 =====================

# ===================== リアルタイムオッズ取得 =====================
def fetch_tansho_odds(race_id: str, timeout: int = 15) -> dict[str, float]:
    """
    netkeibaのオッズAPIから単勝オッズを取得する。
    返り値は {馬番: オッズ} の辞書（例: {"1": 4.2, "2": 18.8}）。
    取得失敗時や発売前は空辞書を返す。
    """
    url = (f"https://race.netkeiba.com/api/api_get_jra_odds.html"
           f"?race_id={race_id}&type=1&action=init")
    try:
        r = SESSION.get(url, timeout=timeout)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        print(f"[WARN] オッズ取得失敗 ({race_id}): {type(e).__name__}: {e}")
        return {}

    # status: "result"=確定 / "middle"=発売中の暫定 / "yoso"=発売前の予想オッズ
    # statusでは弾かず、オッズデータが取れたものをそのまま使う
    status = data.get("status")
    if status == "yoso":
        print(f"[INFO] 発売前のため予想オッズを使用します ({race_id})")

    tansho = data.get("data", {}).get("odds", {}).get("1", {})
    result: dict[str, float] = {}
    for umaban, values in tansho.items():
        # values は [オッズ, 変動値, 人気順] の形式
        try:
            result[str(int(umaban))] = float(values[0])
        except (ValueError, TypeError, IndexError):
            continue  # 発売前は "---.-" などが入るためスキップ
    return result
# ===================== リアルタイムオッズ取得 =====================

def extract_horse_table(html: str) -> list[dict]:
    """
    馬柱(5走)テーブルから
    馬番, 馬名, 性齢, 騎手名,
    前走/2走/3走/4走の(レース名, 場所, コース, 着順,3F)
    を 1頭=1辞書 のリストにして返す
    """
    soup = BeautifulSoup(html, "lxml")

    table = soup.select_one("table.Shutuba_Past5_Table")
    if table is None:
        raise ValueError("Shutuba_Past5_Table が見つかりませんでした")

    rows = table.select("tbody tr.HorseList")

    records = []

    for tr in rows:
        # ───────── 馬番 ─────────
        uma_no = ""
        td_umaban = tr.select_one("td.Waku")
        if not td_umaban:
            # もしクラス名が違う場合の保険
            td_umaban = tr.select_one("td.Umaban")
        if td_umaban:
            uma_no = td_umaban.get_text(strip=True)

        # 馬名（Horse_Info内の Horse02 の a）
        horse_name = ""
        a_horse = tr.select_one("td.Horse_Info div.Horse02 a")
        if a_horse:
            horse_name = a_horse.get_text(strip=True)

        # 性齢（Barei）
        sex_age = ""
        span_barei = tr.select_one("td.Jockey span.Barei")
        if span_barei:
            tmp = span_barei.get_text(strip=True)
            if tmp not in ("性齢、毛色", "勝負服", "騎手"):
                sex_age = tmp

        # 騎手名
        jockey_name = ""
        a_jockey = tr.select_one('td.Jockey a[href*="/jockey/"]')
        if a_jockey:
            jockey_name = a_jockey.get_text(strip=True)

        # 過去走（前走〜5走まで入っている想定）
        past_tds = tr.select("td.Past")

        # 取りたいのは 前走, 2走, 3走, 4走 の4つ
        labels = ["前走", "2走", "3走", "4走"]
        past_data = {}
        for i, label in enumerate(labels):
            if i < len(past_tds):
                race_name, place, course, finish, margin, passing, last3f = parse_past_cell(past_tds[i])
            else:
                race_name, place, course, finish, margin, passing, last3f = "", "", "", "", "", "", ""

            past_data[f"{label}_レース名"] = race_name
            past_data[f"{label}_場所"] = place
            past_data[f"{label}_コース"] = course
            past_data[f"{label}_着順"] = finish
            past_data[f"{label}_着差"] = margin
            past_data[f"{label}_通過順"] = passing
            past_data[f"{label}_３F"] = last3f

        # ───────── 追加項目 ─────────
        # ハンデ（今走の斤量）: td.Jockey 内の数値スパン（例: 58.0）
        handicap = ""
        td_j = tr.select_one("td.Jockey")
        if td_j:
            for sp in td_j.select("span"):
                t = sp.get_text(strip=True)
                if re.fullmatch(r"\d{2}(?:\.\d)?", t):
                    handicap = float(t)
                    break

        # 前走セルから 騎手名・走破タイム・馬体重 を取得
        prev_jockey = ""
        prev_time = ""
        prev_weight = ""
        if past_tds:
            fp = past_tds[0]
            d3 = fp.select_one("div.Data03")
            if d3:
                # 例: "14頭 11番 1人 ルメール 58.0" → 数値始まり以外の最後のトークンが騎手名
                for tok in d3.get_text(" ", strip=True).split():
                    if not re.match(r"^[\d]", tok):
                        prev_jockey = tok
            d5 = fp.select_one("div.Data05")
            if d5:
                mt = re.search(r"\d:\d{2}\.\d", d5.get_text(" ", strip=True))
                if mt:
                    prev_time = mt.group(0)
            d6 = fp.select_one("div.Data06")
            if d6:
                mw = re.search(r"\d{3}\([+\-±0-9]*\)", d6.get_text(" ", strip=True))
                if mw:
                    prev_weight = mw.group(0)

        # 騎手乗替: 今走騎手と前走騎手の比較
        # netkeibaは短縮表記（例: ルメー/ルメール、Ｍデム/Ｍ．デム）のため、
        # 記号を除去したうえで前方一致なら同一騎手とみなす
        jockey_change = ""
        norm = lambda s: re.sub(r"[\s．.・･]", "", s)
        a, b = norm(jockey_name), norm(prev_jockey)
        if a and b and not (a.startswith(b) or b.startswith(a)):
            jockey_change = "乗替"

        # 3着内率: 馬柱にある過去走（最大5走）の着順から算出
        finishes = []
        for td in past_tds[:5]:
            num = td.select_one("div.Data01 span.Num")
            if num:
                t = num.get_text(strip=True)
                if t.isdigit():
                    finishes.append(int(t))
        top3_rate = round(sum(1 for f in finishes if f <= 3) / len(finishes), 3) if finishes else ""

        record = {
            "馬番": uma_no,
            "馬名": horse_name,
            "性齢": sex_age,
            "騎手名": jockey_name,
            "ハンデ": handicap,
            "馬体重": prev_weight,
            "騎手乗替": jockey_change,
            "3着内率": top3_rate,
            "タイム": prev_time,
        }
        record.update(past_data)
        records.append(record)

    # 書き込みはテンプレートの列名と突き合わせるため、列順は問わない
    return records

# ===================== サイトからデータ取得 =====================

# ===================== アウトプットフォルダ作成 =====================
def output_dir() -> Path:
    try:
        base = Path(__file__).resolve().parent
    except NameError:
        base = Path.cwd()
    out = base / "output"
    out.mkdir(parents=True, exist_ok=True)
    return out
# ===================== アウトプットフォルダ作成 =====================

# ===================== シート名安全化 =====================
def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "sheet"
    base = base[:31]
    cand, i = base, 2
    while cand in used:
        suf = f"_{i}"
        cand = (base[:max(0, 31 - len(suf))] + suf)[:31]
        i += 1
    used.add(cand)
    return cand
# ===================== シート名安全化 =====================

# ===================== テンプレートシートへデータ書き込み =====================
def write_df_to_sheet(ws, records: list[dict]):
    """テンプレートの列名とデータの列名を突き合わせて正しい列に書き込む"""
    # テンプレート1行目のヘッダーから 列名→列番号 マッピングを構築
    col_map: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None and not isinstance(cell, MergedCell):
            col_map[str(cell.value)] = cell.column

    # データを2行目から書き込む（列名でマッチング）
    for r_idx, rec in enumerate(records, start=2):
        for col_name, value in rec.items():
            if col_name not in col_map:
                continue
            cell = ws.cell(row=r_idx, column=col_map[col_name])
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = value
# ===================== テンプレートシートへデータ書き込み =====================

def fetch_race(rid: str):
    """1レース分の馬柱・メタ情報・オッズをまとめて取得する（並列実行用）"""
    race_url = f"https://race.netkeiba.com/race/shutuba_past.html?race_id={rid}&rf=shutuba_submenu"
    html = _get_html(race_url)
    _, name, place, rnum, d1, d2 = _extract_race_meta(html)
    records = extract_horse_table(html)

    # リアルタイム単勝オッズをテンプレートの「オッズ」列に反映
    odds_map = fetch_tansho_odds(rid)
    if odds_map:
        for rec in records:
            rec["オッズ"] = odds_map.get(str(rec.get("馬番", "")).strip(), "")

    return name, place, rnum, d1, d2, records


def main():
    t_start = time.time()
    # オプションで WIN5ページのURL上書きも可
    url_arg = sys.argv[1] if len(sys.argv) >= 2 else None

    # WIN5 対象レースの race_id を取得
    race_ids, race_date = pick_win5_ids(url_arg)
    if not race_ids:
        print("対象の WIN5 race_id を取得できませんでした。")
        sys.exit(2)

    # 出力ファイル名
    nowstamp = dt.datetime.now().strftime("%Y%m%d%H%M%S")
    out_dir = output_dir()
    out_xlsx = out_dir / f"Win5軸馬決定_{race_date}_{nowstamp}.xlsx"
    print(f"出力開始: {out_xlsx}")

    # テンプレート読込
    if not TEMPLATE_XLSX.exists():
        print(f"テンプレートが見つかりません: {TEMPLATE_XLSX}")
        sys.exit(3)

    # 全レースを並列取得しつつ、その間にテンプレートを読み込む
    print(f"{len(race_ids)}レースを並列取得中…")
    with ThreadPoolExecutor(max_workers=len(race_ids)) as ex:
        futures = [ex.submit(fetch_race, rid) for rid in race_ids]

        # テンプレートをベースにワークブックを開く（書式・条件付き書式を引き継ぐ）
        wb = load_workbook(TEMPLATE_XLSX)
        template_sheets = wb.worksheets  # 既存5枚シート

        used_sheet_names: set[str] = set()
        errors: list[str] = []
        written = 0

        for idx_r, (rid, future) in enumerate(zip(race_ids, futures)):
            if idx_r >= len(template_sheets):
                print(f"[WARN] テンプレートシートが足りません（{idx_r+1}枚目なし）")
                break
            try:
                name, place, rnum, d1, d2, records = future.result()
                sheet_title = name
                if place and rnum:
                    sheet_title = f"{place}{rnum}_{name}"
                sheet_title = safe_sheet_name(sheet_title, used_sheet_names)

                ws = template_sheets[idx_r]
                ws.title = sheet_title
                write_df_to_sheet(ws, records)

                # レース情報（D25: 発走時刻 / D26: シート名 / D27: 条件・距離）
                ws["D25"] = _parse_race_time(d1)
                ws["D26"] = sheet_title
                ws["D27"] = _parse_course_label(d1, d2)

                print(f"[{written+1}] {sheet_title} に書き込み完了（{len(records)}頭）")
                written += 1
            except Exception as e:
                msg = f"{rid}: {type(e).__name__}: {e}"
                print("[SKIP]", msg)
                errors.append(msg)

    wb.save(out_xlsx)
    print(f"出力完了: {out_xlsx}（{written}/{len(race_ids)}レース書き込み、{time.time() - t_start:.1f}秒）")
    if errors:
        print("エラーがあったレース:")
        for msg in errors:
            print(" -", msg)


if __name__ == "__main__":
    main()
