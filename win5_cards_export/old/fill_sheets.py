# -*- coding: utf-8 -*-
"""
main_horse_decide_sheets.xlsx の WIN 1〜5 シートに
netkeiba の馬柱データを書き込む

使い方:
  python fill_sheets.py                   # 今週のWIN5（idx=0:土曜, 1:日曜）
  python fill_sheets.py <netkeiba_win5_url>
"""
import os
import re
import sys
import time
import shutil
import datetime as dt

import requests
from bs4 import BeautifulSoup
from bs4 import UnicodeDammit
from openpyxl import load_workbook
from pathlib import Path
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager

# ===================== 定数 =====================
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36"),
    "Referer": "https://www.netkeiba.com/",
    "Accept-Language": "ja,en;q=0.9",
}
WIN5_IDX = 0  # 土曜日は0、日曜日は1
PC_URL = f"https://race.netkeiba.com/top/win5.html?idx={WIN5_IDX}"
RACE_ID_RE    = re.compile(r"race_id=(\d{12})")
PLACE_PATTERN = re.compile(r"(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)")

# テンプレートファイル（スクリプトと同じフォルダ）
TEMPLATE_XLSX = Path(__file__).resolve().parent / "main_horse_decide_sheets.xlsx"

# ── 列マッピング（1ベース列番号） ──────────────────────────
# A(1)=予想[手動], B(2)=馬番, C(3)=オッズ[手動],
# D(4)=馬名, E(5)=性齢, F(6)=騎手名
# G(7)〜AH(34)=前走〜4走データ（各7列）
# AI(35)〜AO(41)=計算式, AP(42)=脚質[手動]
DATA_COL_MAP: dict[str, int] = {
    "馬番":         2,
    "オッズ":       3,
    "馬名":         4,
    "性齢":         5,
    "騎手名":       6,
    "前走_レース名": 7,
    "前走_場所":    8,
    "前走_コース":  9,
    "前走_着順":    10,
    "前走_着差":    11,
    "前走_通過順":  12,
    "前走_３F":     13,
    "2走_レース名": 14,
    "2走_場所":     15,
    "2走_コース":   16,
    "2走_着順":     17,
    "2走_着差":     18,
    "2走_通過順":   19,
    "2走_３F":      20,
    "3走_レース名": 21,
    "3走_場所":     22,
    "3走_コース":   23,
    "3走_着順":     24,
    "3走_着差":     25,
    "3走_通過順":   26,
    "3走_３F":      27,
    "4走_レース名": 28,
    "4走_場所":     29,
    "4走_コース":   30,
    "4走_着順":     31,
    "4走_着差":     32,
    "4走_通過順":   33,
    "4走_３F":      34,
}
# 計算式列（AI=35〜AO=41）は上書きしない
FORMULA_COL_START = 35
FORMULA_COL_END   = 41


# ===================== HTTP セッション =====================
def _build_session() -> requests.Session:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    s = requests.Session()
    s.headers.update(HEADERS)
    retry = Retry(
        total=3, backoff_factor=0.3,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://",  HTTPAdapter(max_retries=retry))
    return s

SESSION = _build_session()


def _get_html(url: str, timeout: int = 15) -> str:
    r = SESSION.get(url, timeout=timeout)
    r.raise_for_status()
    dammit = UnicodeDammit(r.content, is_html=True)
    return dammit.unicode_markup or r.content.decode("utf-8", errors="replace")


# ===================== WIN5 race_id 取得 =====================
def _extract_ids(html: str) -> list[str]:
    soup = BeautifulSoup(html, "html.parser")
    ids, seen = [], set()
    for a in soup.find_all("a", href=True):
        m = RACE_ID_RE.search(a["href"])
        if not m:
            continue
        rid = m.group(1)
        if rid not in seen:
            seen.add(rid)
            ids.append(rid)
    return ids


def _race_date_from_html(html: str) -> str:
    soup = BeautifulSoup(html, "html.parser")
    ids = _extract_ids(html)
    year = ids[0][:4] if ids else str(dt.datetime.now().year)

    for sel in (".RaceList_Date dl.Win5_Date dd.Active", ".RaceList_Date dd.Active"):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            m = re.search(r"(\d{1,2})月\s*(\d{1,2})日", txt)
            if m:
                return f"{int(year):04d}{int(m.group(1)):02d}{int(m.group(2)):02d}"
    return ""


def pick_win5_ids(url: str | None = None) -> tuple[list[str], str]:
    target = url or PC_URL
    try:
        html = _get_html(target)
        ids  = _extract_ids(html)
        date = _race_date_from_html(html)
    except Exception as e:
        print(f"[ERROR] WIN5ページ取得失敗: {e}")
        return [], ""
    return ids[:5], date


# ===================== 馬柱パース =====================
def _parse_past_td(td) -> tuple[str, str, str, str, str, str, str]:
    """<td class="Past"> から (レース名, 場所, コース, 着順, 着差, 通過順, 3F) を返す"""
    if td is None:
        return ("",) * 7

    race_name = ""
    a_tag = td.select_one("div.Data02 a")
    if a_tag:
        t = a_tag.find(string=True, recursive=False)
        if t:
            race_name = t.strip()

    place = ""
    span = td.select_one("div.Data01 span:not(.Num)")
    if span:
        parts = span.get_text(strip=True).split()
        if len(parts) >= 2:
            place = parts[-1]

    course = ""
    d05 = td.select_one("div.Data05")
    if d05:
        parts = d05.get_text(" ", strip=True).split()
        if parts:
            course = parts[0]

    finish = ""
    num = td.select_one("div.Data01 span.Num")
    if num:
        finish = num.get_text(strip=True)

    margin = ""
    d07 = td.select_one("div.Data07")
    if d07:
        m = re.search(r"\(([^)]+)\)", d07.get_text(" ", strip=True))
        if m:
            margin = m.group(1).strip()

    passing = last3f = ""
    d06 = td.select_one("div.Data06")
    if d06:
        t = d06.get_text(" ", strip=True)
        if t:
            first = t.split()[0]
            if re.fullmatch(r"\d+(?:-\d+)+", first):
                passing = first
        m = re.search(r"\(([\d\.]+)\)", t)
        if m:
            last3f = m.group(1)

    return race_name, place, course, finish, margin, passing, last3f


class _LazyBrowser:
    """必要な時だけ起動し、プロセスは使い回す。"""
    def __init__(self):
        self._driver = None

    def _new_driver(self):
        os.environ["WDM_LOG"] = "0"
        os.environ["WDM_PRINT_FIRST_LINE"] = "False"
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        options.add_argument("--blink-settings=imagesEnabled=false")
        options.add_argument("--lang=ja-JP")
        options.add_argument("--log-level=3")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        service = ChromeService(
            ChromeDriverManager().install(),
            log_output=open(os.devnull, "w", encoding="utf-8", errors="ignore"),
        )
        d = webdriver.Chrome(service=service, options=options)
        d.set_page_load_timeout(45)
        return d

    @property
    def driver(self):
        if self._driver is None:
            self._driver = self._new_driver()
        return self._driver

    def get_html(self, url: str, wait_selector: str, timeout: int = 20) -> str:
        d = self.driver
        try:
            d.get(url)
        except TimeoutException:
            pass
        t0 = time.time()
        while time.time() - t0 < timeout:
            try:
                if d.find_elements("css selector", wait_selector):
                    break
            except Exception:
                pass
            time.sleep(0.5)
        return d.page_source

    def close(self):
        if self._driver:
            try:
                self._driver.quit()
            except Exception:
                pass
            self._driver = None


BROWSER = _LazyBrowser()


def _parse_odds_html(html: str) -> dict[str, float]:
    """レンダリング済みHTMLから {馬番: オッズ} を抽出する"""
    soup = BeautifulSoup(html, "html.parser")
    odds_map: dict[str, float] = {}
    # テーブル構造: 人気 | 馬番 | 印 | 馬名 | 予想オッズ
    for tr in soup.select("tr"):
        cells = tr.find_all("td")
        if len(cells) < 4:
            continue
        texts = [td.get_text(strip=True) for td in cells]
        if not (texts[0].isdigit() and texts[1].isdigit()):
            continue
        umaban = texts[1]
        for t in texts[2:]:
            t_clean = t.replace(",", "")
            if re.fullmatch(r"\d+\.\d+", t_clean):
                try:
                    odds_map[umaban] = float(t_clean)
                    break
                except ValueError:
                    pass
    return odds_map


def fetch_odds(race_id: str) -> dict[str, float]:
    """
    Selenium でオッズページをレンダリングし {馬番(str): 単勝オッズ(float)} を返す。
    取得できない場合は空の辞書を返す。
    """
    url = (f"https://race.netkeiba.com/odds/index.html"
           f"?race_id={race_id}&rf=race_submenu")
    try:
        # オッズ数値が入るセルが描画されるまで待つ
        html = BROWSER.get_html(url, wait_selector="table.Odds_Table tbody tr td")
        return _parse_odds_html(html)
    except Exception as e:
        print(f"  [WARN] オッズSelenium取得失敗: {e}")
        return {}


def _extract_place_rnum(html: str) -> tuple[str, str]:
    """HTMLから会場名・レース番号を取得する（例: '阪神', '9R'）"""
    soup = BeautifulSoup(html, "html.parser")
    place = ""
    d2 = soup.select_one(".RaceData02")
    if d2:
        m = PLACE_PATTERN.search(d2.get_text(" ", strip=True))
        if m:
            place = m.group(1)
    rnum = ""
    rnum_el = soup.select_one(".RaceNum")
    if rnum_el:
        m = re.search(r"(\d+)R", rnum_el.get_text(strip=True), re.I)
        if m:
            rnum = m.group(1) + "R"
    return place, rnum


def fetch_horse_rows(race_id: str) -> tuple[list[dict], str, str]:
    """
    shutuba_past ページから馬ごとのデータと会場・R番を返す
    戻り値: (records, place, rnum)  例: ([...], '阪神', '9R')
    """
    url  = (f"https://race.netkeiba.com/race/shutuba_past.html"
            f"?race_id={race_id}&rf=shutuba_submenu")
    html = _get_html(url)
    place, rnum = _extract_place_rnum(html)
    soup = BeautifulSoup(html, "html.parser")

    table = soup.select_one("table.Shutuba_Past5_Table")
    if table is None:
        raise ValueError(f"Shutuba_Past5_Table が見つかりません (race_id={race_id})")

    records = []
    for tr in table.select("tbody tr.HorseList"):
        # 馬番
        uma_no = ""
        for cls in ("td.Waku", "td.Umaban"):
            el = tr.select_one(cls)
            if el:
                uma_no = el.get_text(strip=True)
                break

        # 馬名
        horse_name = ""
        a = tr.select_one("td.Horse_Info div.Horse02 a")
        if a:
            horse_name = a.get_text(strip=True)

        # 性齢
        sex_age = ""
        span = tr.select_one("td.Jockey span.Barei")
        if span:
            tmp = span.get_text(strip=True)
            if tmp not in ("性齢、毛色", "勝負服", "騎手"):
                sex_age = tmp

        # 騎手名
        jockey = ""
        a_j = tr.select_one('td.Jockey a[href*="/jockey/"]')
        if a_j:
            jockey = a_j.get_text(strip=True)

        past_tds = tr.select("td.Past")
        rec: dict = {"馬番": uma_no, "馬名": horse_name, "性齢": sex_age, "騎手名": jockey}
        for i, label in enumerate(("前走", "2走", "3走", "4走")):
            td = past_tds[i] if i < len(past_tds) else None
            r_name, p, crs, fin, mrg, passing, f3 = _parse_past_td(td)
            rec[f"{label}_レース名"] = r_name
            rec[f"{label}_場所"]     = p
            rec[f"{label}_コース"]   = crs
            rec[f"{label}_着順"]     = fin
            rec[f"{label}_着差"]     = mrg
            rec[f"{label}_通過順"]   = passing
            rec[f"{label}_３F"]      = f3
        records.append(rec)

    # 単勝オッズを取得して各レコードに追加
    print(f"  オッズ取得中…")
    odds_map = fetch_odds(race_id)
    for rec in records:
        rec["オッズ"] = odds_map.get(str(rec.get("馬番", "")), None)
    if odds_map:
        print(f"  オッズ取得完了 ({len(odds_map)} 頭分)")
    else:
        print(f"  オッズ取得できませんでした（レース前など）")

    return records, place, rnum


# ===================== 計算式生成 =====================
def _row_formulas(row: int) -> dict[int, str]:
    """行番号 row に対して計算式辞書 {列番号: 数式} を返す"""
    r = row
    return {
        35: f"=MIN(AH{r},AA{r},T{r},M{r})",          # AI: 最速3F
        36: f"=SUM(K{r},R{r},Y{r},AF{r})",            # AJ: 着差合計
        37: (f"=COUNTIF(AE{r},1)+COUNTIF(X{r},1)"
             f"+COUNTIF(Q{r},1)+COUNTIF(J{r},1)"),    # AK: 1着
        38: (f"=COUNTIF(AE{r},2)+COUNTIF(X{r},2)"
             f"+COUNTIF(Q{r},2)+COUNTIF(J{r},2)"),    # AL: 2着
        39: (f"=COUNTIF(AE{r},3)+COUNTIF(X{r},3)"
             f"+COUNTIF(Q{r},3)+COUNTIF(J{r},3)"),    # AM: 3着
        40: (f"=COUNTIF(AE{r},4)+COUNTIF(X{r},4)"
             f"+COUNTIF(Q{r},4)+COUNTIF(J{r},4)"),    # AN: 4着
        41: f"=SUM(AK{r}:AN{r})",                     # AO: 合計
    }


# ===================== シートへの書き込み =====================
def fill_worksheet(ws, records: list[dict]) -> None:
    """
    ワークシート ws の 2行目以降にデータを書き込む。
    - 列 B(2), D(4)〜AH(34) を上書き
    - 列 A(1)=予想, C(3)=オッズ, AP(42)=脚質 は触れない
    - 列 AI(35)〜AO(41) は計算式を補完（既存があれば保持）
    """
    # 既存データ行をクリア（数式列・手動入力列は除く）
    CLEAR_COLS = set(range(2, FORMULA_COL_START))  # B〜AH（C=オッズ含む）
    for row_idx in range(2, ws.max_row + 1):
        for col in CLEAR_COLS:
            ws.cell(row=row_idx, column=col).value = None

    # データ書き込み + 計算式補完
    for i, rec in enumerate(records):
        row = i + 2  # 1行目はヘッダ
        for field, col in DATA_COL_MAP.items():
            val = rec.get(field, "")
            # 数値変換を試みる（着順・着差・3Fなど）
            if val != "" and val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row, column=col).value = val if val != "" else None

        # 計算式列：既存の値がなければ書き込む
        for col, formula in _row_formulas(row).items():
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = formula


# ===================== メイン =====================
def main() -> None:
    url_arg = sys.argv[1] if len(sys.argv) >= 2 else None

    # テンプレート確認
    if not TEMPLATE_XLSX.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE_XLSX}")
        sys.exit(1)

    # WIN5 race_id 取得
    race_ids, win5_date = pick_win5_ids(url_arg)
    if not race_ids:
        print("[ERROR] WIN5 の race_id を取得できませんでした。")
        sys.exit(2)

    print(f"WIN5 対象レース ({win5_date}): {race_ids}")

    # 出力ファイル（テンプレートをコピー）
    nowstamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = TEMPLATE_XLSX.parent / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path  = out_dir / f"main_horse_decide_sheets_{nowstamp}.xlsx"
    shutil.copy(TEMPLATE_XLSX, out_path)
    print(f"出力ファイル: {out_path}")

    wb = load_workbook(out_path)
    sheet_names = ["WIN 1", "WIN 2", "WIN 3", "WIN 4", "WIN 5"]

    errors: list[str] = []
    for i, race_id in enumerate(race_ids):
        sheet_name = sheet_names[i] if i < len(sheet_names) else f"WIN {i+1}"
        print(f"[{i+1}/5] {sheet_name}  race_id={race_id} を取得中…")
        try:
            records, place, rnum = fetch_horse_rows(race_id)
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シート '{sheet_name}' が見つかりません")
            ws = wb[sheet_name]
            fill_worksheet(ws, records)
            # シート名を「WIN X_阪神9R」形式に変更
            if place and rnum:
                ws.title = f"{sheet_name}_{place}{rnum}"
            print(f"[{i+1}/5] {ws.title}  {len(records)} 頭 書き込み完了")
        except Exception as e:
            msg = f"{sheet_name} ({race_id}): {type(e).__name__}: {e}"
            print(f"[SKIP] {msg}")
            errors.append(msg)

    wb.save(out_path)
    BROWSER.close()
    print(f"\n保存完了: {out_path}")
    if errors:
        print(f"\nスキップ ({len(errors)} 件):")
        for m in errors:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
