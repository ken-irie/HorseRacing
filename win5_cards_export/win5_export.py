# -*- coding: utf-8 -*-
"""
WIN5 馬柱データ＋単勝オッズを xlsx に書き込む

使い方:
  python win5_export.py                   # 今週のWIN5（idx=0:土曜, 1:日曜）
  python win5_export.py <netkeiba_win5_url>
"""
import os
import re
import sys
import json
import shutil
import datetime as dt
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager


# ===================== 定数 =====================
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/124.0.0.0 Safari/537.36"),
    "Referer": "https://www.netkeiba.com/",
    "Accept-Language": "ja,en;q=0.9",
}
WIN5_IDX = 0  # 土曜日は0、日曜日は1
PC_URL = f"https://race.netkeiba.com/top/win5.html?idx={WIN5_IDX}"
RACE_ID_RE    = re.compile(r"race_id=(\d{12})")
PLACE_PATTERN = re.compile(r"(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)")

TEMPLATE_XLSX = Path(__file__).resolve().parent / "main_horse_decide_sheets.xlsx"

# 列マッピング（1ベース）
# A(1)=予想[手動], B(2)=馬番, C(3)=オッズ, D(4)=馬名, E(5)=性齢, F(6)=騎手名
# G(7)〜AH(34)=前走〜4走データ, AI(35)〜AO(41)=計算式, AP(42)=脚質[手動]
DATA_COL_MAP: dict[str, int] = {
    "馬番":         2,
    "馬名":         4,
    "性齢":         5,
    "騎手名":       6,
    "前走_レース名": 7,
    "前走_場所":    8,
    "前走_コース":  9,
    "前走_着順":    10,
    "前走_着差":    11,
    "前走_通過順":  12,
    "前走_３F":     13,
    "2走_レース名": 14,
    "2走_場所":     15,
    "2走_コース":   16,
    "2走_着順":     17,
    "2走_着差":     18,
    "2走_通過順":   19,
    "2走_３F":      20,
    "3走_レース名": 21,
    "3走_場所":     22,
    "3走_コース":   23,
    "3走_着順":     24,
    "3走_着差":     25,
    "3走_通過順":   26,
    "3走_３F":      27,
    "4走_レース名": 28,
    "4走_場所":     29,
    "4走_コース":   30,
    "4走_着順":     31,
    "4走_着差":     32,
    "4走_通過順":   33,
    "4走_３F":      34,
}
FORMULA_COL_START = 35
FORMULA_COL_END   = 41
UMABAN_COL        = 2
ODDS_COL          = 3


# ===================== HTTP セッション =====================
def _build_session() -> requests.Session:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    s = requests.Session()
    s.headers.update(HEADERS)
    retry = Retry(
        total=3, backoff_factor=0.3,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",),
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.mount("http://",  HTTPAdapter(max_retries=retry))
    return s

SESSION = _build_session()


def _get_html(url: str, timeout: int = 15) -> str:
    r = SESSION.get(url, timeout=timeout)
    r.raise_for_status()
    dammit = UnicodeDammit(r.content, is_html=True)
    return dammit.unicode_markup or r.content.decode("utf-8", errors="replace")


# ===================== WIN5 race_id 取得 =====================
def _extract_ids_from_soup(soup: BeautifulSoup) -> list[str]:
    ids, seen = [], set()
    for a in soup.find_all("a", href=True):
        m = RACE_ID_RE.search(a["href"])
        if not m:
            continue
        rid = m.group(1)
        if rid not in seen:
            seen.add(rid)
            ids.append(rid)
    return ids


def _race_date_from_soup(soup: BeautifulSoup, ids: list[str]) -> str:
    year = ids[0][:4] if ids else str(dt.datetime.now().year)
    for sel in (".RaceList_Date dl.Win5_Date dd.Active", ".RaceList_Date dd.Active"):
        node = soup.select_one(sel)
        if node:
            txt = node.get_text(" ", strip=True)
            m = re.search(r"(\d{1,2})月\s*(\d{1,2})日", txt)
            if m:
                return f"{int(year):04d}{int(m.group(1)):02d}{int(m.group(2)):02d}"
    return ""


def pick_win5_ids(url: str | None = None) -> tuple[list[str], str]:
    try:
        html = _get_html(url or PC_URL)
        soup = BeautifulSoup(html, "html.parser")
        ids  = _extract_ids_from_soup(soup)
        date = _race_date_from_soup(soup, ids)
    except Exception as e:
        print(f"[ERROR] WIN5ページ取得失敗: {e}")
        return [], ""
    return ids[:5], date


# ===================== 馬柱パース =====================
def _parse_past_td(td) -> tuple[str, str, str, str, str, str, str]:
    if td is None:
        return ("",) * 7

    race_name = ""
    a_tag = td.select_one("div.Data02 a")
    if a_tag:
        t = a_tag.find(string=True, recursive=False)
        if t:
            race_name = t.strip()

    place = ""
    span = td.select_one("div.Data01 span:not(.Num)")
    if span:
        parts = span.get_text(strip=True).split()
        if len(parts) >= 2:
            place = parts[-1]

    course = ""
    d05 = td.select_one("div.Data05")
    if d05:
        parts = d05.get_text(" ", strip=True).split()
        if parts:
            course = parts[0]

    finish = ""
    num = td.select_one("div.Data01 span.Num")
    if num:
        finish = num.get_text(strip=True)

    margin = ""
    d07 = td.select_one("div.Data07")
    if d07:
        m = re.search(r"\(([^)]+)\)", d07.get_text(" ", strip=True))
        if m:
            margin = m.group(1).strip()

    passing = last3f = ""
    d06 = td.select_one("div.Data06")
    if d06:
        t = d06.get_text(" ", strip=True)
        if t:
            first = t.split()[0]
            if re.fullmatch(r"\d+(?:-\d+)+", first):
                passing = first
        m = re.search(r"\(([\d\.]+)\)", t)
        if m:
            last3f = m.group(1)

    return race_name, place, course, finish, margin, passing, last3f


def _extract_place_rnum_from_soup(soup: BeautifulSoup) -> tuple[str, str]:
    place = ""
    d2 = soup.select_one(".RaceData02")
    if d2:
        m = PLACE_PATTERN.search(d2.get_text(" ", strip=True))
        if m:
            place = m.group(1)
    rnum = ""
    rnum_el = soup.select_one(".RaceNum")
    if rnum_el:
        m = re.search(r"(\d+)R", rnum_el.get_text(strip=True), re.I)
        if m:
            rnum = m.group(1) + "R"
    return place, rnum


def fetch_horse_rows(race_id: str) -> tuple[list[dict], str, str]:
    url  = (f"https://race.netkeiba.com/race/shutuba_past.html"
            f"?race_id={race_id}&rf=shutuba_submenu")
    html = _get_html(url)
    soup = BeautifulSoup(html, "html.parser")
    place, rnum = _extract_place_rnum_from_soup(soup)

    table = soup.select_one("table.Shutuba_Past5_Table")
    if table is None:
        raise ValueError(f"Shutuba_Past5_Table が見つかりません (race_id={race_id})")

    records = []
    for tr in table.select("tbody tr.HorseList"):
        uma_no = ""
        for cls in ("td.Waku", "td.Umaban"):
            el = tr.select_one(cls)
            if el:
                uma_no = el.get_text(strip=True)
                break

        horse_name = ""
        a = tr.select_one("td.Horse_Info div.Horse02 a")
        if a:
            horse_name = a.get_text(strip=True)

        sex_age = ""
        span = tr.select_one("td.Jockey span.Barei")
        if span:
            tmp = span.get_text(strip=True)
            if tmp not in ("性齢、毛色", "勝負服", "騎手"):
                sex_age = tmp

        jockey = ""
        a_j = tr.select_one('td.Jockey a[href*="/jockey/"]')
        if a_j:
            jockey = a_j.get_text(strip=True)

        past_tds = tr.select("td.Past")
        rec: dict = {"馬番": uma_no, "馬名": horse_name, "性齢": sex_age, "騎手名": jockey}
        for i, label in enumerate(("前走", "2走", "3走", "4走")):
            td = past_tds[i] if i < len(past_tds) else None
            r_name, p, crs, fin, mrg, passing, f3 = _parse_past_td(td)
            rec[f"{label}_レース名"] = r_name
            rec[f"{label}_場所"]     = p
            rec[f"{label}_コース"]   = crs
            rec[f"{label}_着順"]     = fin
            rec[f"{label}_着差"]     = mrg
            rec[f"{label}_通過順"]   = passing
            rec[f"{label}_３F"]      = f3
        records.append(rec)

    return records, place, rnum


# ===================== 計算式生成 =====================
def _row_formulas(row: int) -> dict[int, str]:
    r = row
    return {
        35: f"=MIN(AH{r},AA{r},T{r},M{r})",
        36: f"=SUM(K{r},R{r},Y{r},AF{r})",
        37: (f"=COUNTIF(AE{r},1)+COUNTIF(X{r},1)"
             f"+COUNTIF(Q{r},1)+COUNTIF(J{r},1)"),
        38: (f"=COUNTIF(AE{r},2)+COUNTIF(X{r},2)"
             f"+COUNTIF(Q{r},2)+COUNTIF(J{r},2)"),
        39: (f"=COUNTIF(AE{r},3)+COUNTIF(X{r},3)"
             f"+COUNTIF(Q{r},3)+COUNTIF(J{r},3)"),
        40: (f"=COUNTIF(AE{r},4)+COUNTIF(X{r},4)"
             f"+COUNTIF(Q{r},4)+COUNTIF(J{r},4)"),
        41: f"=SUM(AK{r}:AN{r})",
    }


# ===================== シートへの書き込み =====================
def fill_worksheet(ws, records: list[dict]) -> None:
    CLEAR_COLS = [c for c in range(2, FORMULA_COL_START) if c != ODDS_COL]
    clear_end = max(ws.max_row, len(records) + 1)
    for row_idx in range(2, clear_end + 1):
        for col in CLEAR_COLS:
            ws.cell(row=row_idx, column=col).value = None

    for i, rec in enumerate(records):
        row = i + 2
        for field, col in DATA_COL_MAP.items():
            val = rec.get(field, "")
            if val != "" and val is not None:
                try:
                    val = float(val)
                    if val == int(val):
                        val = int(val)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row, column=col).value = val if val != "" else None

        for col, formula in _row_formulas(row).items():
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = formula


# ===================== Selenium（オッズ取得） =====================
class _LazyBrowser:
    def __init__(self):
        self._driver = None

    def _new_driver(self):
        os.environ["WDM_LOG"] = "0"
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        )
        options.add_argument("--log-level=3")
        options.add_experimental_option("excludeSwitches", ["enable-logging", "enable-automation"])
        options.add_experimental_option("useAutomationExtension", False)
        service = ChromeService(
            ChromeDriverManager().install(),
            log_output=open(os.devnull, "w", encoding="utf-8", errors="ignore"),
        )
        d = webdriver.Chrome(service=service, options=options)
        d.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        })
        d.set_page_load_timeout(45)
        return d

    @property
    def driver(self):
        if self._driver is None:
            self._driver = self._new_driver()
        return self._driver

    def close(self):
        if self._driver:
            try:
                self._driver.quit()
            except Exception:
                pass
            self._driver = None


BROWSER = _LazyBrowser()


def fetch_odds(race_id: str) -> dict[str, float]:
    """
    shutuba.html を Selenium でレンダリングし、
    id="odds-1_<馬番>" 要素から {馬番(str): 単勝オッズ(float)} を返す。
    """
    url = f"https://race.netkeiba.com/race/shutuba.html?race_id={race_id}"
    d = BROWSER.driver
    try:
        d.get(url)
    except TimeoutException:
        pass

    try:
        WebDriverWait(d, 30).until(
            lambda drv: (
                elems := drv.find_elements(By.CSS_SELECTOR, '[id^="odds-1_"]')
            ) and any(
                e.text.strip() not in ("", "---.-", "---")
                for e in elems
            )
        )
    except TimeoutException:
        return {}

    odds_map: dict[str, float] = {}
    for el in d.find_elements(By.CSS_SELECTOR, '[id^="odds-1_"]'):
        eid    = el.get_attribute("id")
        umaban = str(int(eid.split("_")[-1]))
        try:
            odds_map[umaban] = float(el.text.strip())
        except ValueError:
            pass
    return odds_map


def fill_odds(ws, odds_map: dict[str, float]) -> int:
    """ワークシートの C列(オッズ)に odds_map を書き込む。書き込み頭数を返す。"""
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        umaban_cell = row[UMABAN_COL - 1]
        odds_cell   = row[ODDS_COL - 1]
        if umaban_cell.value is None:
            continue
        raw    = umaban_cell.value
        umaban = str(int(raw)) if isinstance(raw, float) else str(raw).strip()
        odds   = odds_map.get(umaban)
        if odds is not None:
            odds_cell.value = odds
            count += 1
    return count


# ===================== メイン =====================
def main() -> None:
    url_arg = sys.argv[1] if len(sys.argv) >= 2 else None

    if not TEMPLATE_XLSX.exists():
        print(f"[ERROR] テンプレートが見つかりません: {TEMPLATE_XLSX}")
        sys.exit(1)

    # WIN5 race_id 取得
    race_ids, win5_date = pick_win5_ids(url_arg)
    if not race_ids:
        print("[ERROR] WIN5 の race_id を取得できませんでした。")
        sys.exit(2)

    print(f"WIN5 対象レース ({win5_date}): {race_ids}")

    # 出力ファイル（テンプレートをコピー）
    nowstamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir  = TEMPLATE_XLSX.parent / "output"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"main_horse_decide_sheets_{nowstamp}.xlsx"
    shutil.copy(TEMPLATE_XLSX, out_path)

    # 馬柱HTML を並列フェッチ
    print("馬柱データを並列取得中…")
    horse_results: dict[int, tuple[list[dict], str, str] | Exception] = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(fetch_horse_rows, rid): i
                   for i, rid in enumerate(race_ids)}
        for future in as_completed(futures):
            idx = futures[future]
            try:
                horse_results[idx] = future.result()
            except Exception as e:
                horse_results[idx] = e

    # シート書き込み
    wb = load_workbook(out_path)
    sheet_names_base = ["WIN 1", "WIN 2", "WIN 3", "WIN 4", "WIN 5"]
    actual_sheet_names: list[str] = []
    errors: list[str] = []

    for i, race_id in enumerate(race_ids):
        sheet_name = sheet_names_base[i] if i < len(sheet_names_base) else f"WIN {i+1}"
        result = horse_results.get(i)

        if isinstance(result, Exception):
            msg = f"{sheet_name} ({race_id}): {type(result).__name__}: {result}"
            print(f"[SKIP] {msg}")
            errors.append(msg)
            actual_sheet_names.append(sheet_name)
            continue

        records, place, rnum = result
        try:
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"シート '{sheet_name}' が見つかりません")
            ws = wb[sheet_name]
            fill_worksheet(ws, records)
            if place and rnum:
                ws.title = f"{sheet_name}_{place}{rnum}"
            actual_sheet_names.append(ws.title)
            print(f"[{i+1}/5] {ws.title}  {len(records)} 頭 書き込み完了")
        except Exception as e:
            msg = f"{sheet_name} ({race_id}): {type(e).__name__}: {e}"
            print(f"[SKIP] {msg}")
            errors.append(msg)
            actual_sheet_names.append(sheet_name)

    wb.save(out_path)
    print(f"\n馬柱保存完了: {out_path}")

    # オッズ取得・書き込み
    print("\nオッズ取得中（Selenium）…")
    wb2 = load_workbook(out_path)
    for i, (race_id, sheet_name) in enumerate(zip(race_ids, actual_sheet_names)):
        if sheet_name not in wb2.sheetnames:
            print(f"[SKIP] シート '{sheet_name}' が見つかりません")
            errors.append(f"{sheet_name}: シートなし")
            continue

        print(f"[{i+1}/{len(race_ids)}] {sheet_name}  race_id={race_id} オッズ取得中…")
        odds_map = fetch_odds(race_id)

        if not odds_map:
            print(f"  オッズ取得できませんでした（レース前など）")
            errors.append(f"{sheet_name}: オッズ取得失敗")
            continue

        count = fill_odds(wb2[sheet_name], odds_map)
        print(f"  オッズ書き込み完了 ({count} 頭分)")

    wb2.save(out_path)
    BROWSER.close()
    print(f"\n保存完了: {out_path}")

    if errors:
        print(f"\n警告 ({len(errors)} 件):")
        for m in errors:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
