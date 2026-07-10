# -*- coding: utf-8 -*-
import os
import re
import sys
import time
import datetime as dt
import pandas as pd
import requests

from concurrent.futures import ThreadPoolExecutor
from io import StringIO
from pathlib import Path
from bs4 import BeautifulSoup, UnicodeDammit
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# ===================== 定数 =====================
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/120.0.0.0 Safari/537.36"),
    "Referer": "https://www.netkeiba.com/",
    "Accept-Language": "ja,en;q=0.9",
}
# 土曜日はidx=0、日曜日はidx=1（launcher.pyからは環境変数WIN5_IDXで上書きされる）
try:
    idx = int(os.environ.get("WIN5_IDX", "1"))
except ValueError:
    idx = 1
PC_URL = f"https://race.netkeiba.com/top/win5.html?idx={idx}"
SP_URL = "https://race.sp.netkeiba.com/?pid=win5&date={date}"  # YYYYMMDD
SHUTUBA_URL = "https://race.netkeiba.com/race/shutuba.html?race_id={rid}"
# 出馬表ページのJSが叩いている単勝オッズAPI（type=1が単勝）
ODDS_API_URL = "https://race.netkeiba.com/api/api_get_jra_odds.html?race_id={rid}&type=1&action=update"
RACE_ID_RE = re.compile(r"race_id=(\d{12})")

# テンプレートファイル
TEMPLATE_XLSX = Path(__file__).resolve().with_name("race_cards.xlsx")
# オッズデータ入力シートのWIN別セクション開始列（B=2, N=14, Z=26, AL=38, AX=50）
WIN_SECTION_COLS = [2, 14, 26, 38, 50]
# セクション内のデータ列オフセット（数式列の差・人気順はスキップ）
DATA_COL_OFFSETS = {"馬番": 2, "オッズ": 3, "馬名": 4, "性齢": 5, "斤量": 6, "騎手名": 7}
DATA_START_ROW = 8
DATA_END_ROW   = 25
TIME_ROW       = 4
RACE_NAME_ROW  = 5
COURSE_ROW     = 6

# ===================== 高速化：HTTPセッション =====================
def build_session() -> requests.Session:
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry

    s = requests.Session()
    s.headers.update(HEADERS)
    retry = Retry(
        total=3, backoff_factor=0.3,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET",)
    )
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    return s

SESSION = build_session()

# ===================== HTMLユーティリティ =====================
def _decode_html_bytes(b: bytes, fallback: str = "utf-8") -> str:
    dammit = UnicodeDammit(b, is_html=True)
    if dammit.unicode_markup:
        return dammit.unicode_markup
    return b.decode(fallback, errors="replace")

def _get_html(url: str, timeout: int = 15) -> str:
    r = SESSION.get(url, timeout=timeout)
    r.raise_for_status()
    return _decode_html_bytes(r.content)

# ===================== Selenium（必要時のみ） =====================
class LazyBrowser:
    """必要な時だけ起動し、プロセスは使い回す。selenium系のimportも初回起動まで遅延させる。"""
    def __init__(self):
        self._driver = None

    def _new_driver(self):
        from selenium import webdriver
        from selenium.webdriver.chrome.service import Service as ChromeService
        from webdriver_manager.chrome import ChromeDriverManager

        os.environ["WDM_LOG"] = "0"
        os.environ["WDM_PRINT_FIRST_LINE"] = "False"
        options = webdriver.ChromeOptions()
        options.add_argument("--headless=new")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--disable-webgl")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--no-sandbox")
        options.add_argument("--blink-settings=imagesEnabled=false")
        options.add_argument("--lang=ja-JP")
        options.add_argument("--remote-debugging-pipe")  # DevToolsログ抑制
        options.add_experimental_option("excludeSwitches", ["enable-logging"])
        options.add_argument("--log-level=3")
        options.add_argument("--silent")
        service = ChromeService(
            ChromeDriverManager().install(),
            log_output=open(os.devnull, "w", encoding="utf-8", errors="ignore")
        )
        d = webdriver.Chrome(service=service, options=options)
        d.set_page_load_timeout(45)
        d.set_script_timeout(45)
        return d

    @property
    def driver(self):
        if self._driver is None:
            self._driver = self._new_driver()
        return self._driver

    def get_rendered_html(self, url: str, wait_css: str = None, hard_timeout: int = 25, wait_odds: bool = False) -> str:
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.common.exceptions import TimeoutException

        d = self.driver
        try:
            try:
                d.get(url)
            except TimeoutException:
                pass

            t0 = time.time()
            while time.time() - t0 < min(8, hard_timeout):  # DOMContentLoaded 相当を短めに
                if d.execute_script("return document.readyState") in ("interactive", "complete"):
                    break
                time.sleep(0.2)

            if wait_css:
                WebDriverWait(d, hard_timeout).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, wait_css))
                )

            if wait_odds:
                def odds_ready(drv):
                    try:
                        return drv.execute_script("""
                            const nodes = document.querySelectorAll('td.Popular, td.Odds, .Popular, .Odds');
                            for (const n of nodes) {
                              const t=(n.textContent||'').trim();
                              if (/^\\d+(?:\\.\\d+)?(?:\\s*倍)?$/.test(t)) return true;
                            }
                            return false;
                        """)
                    except Exception:
                        return False
                WebDriverWait(d, hard_timeout).until(lambda drv: odds_ready(drv))

            if d.execute_script("return document.readyState") != "complete":
                d.execute_script("window.stop();")
            return d.page_source
        except Exception:
            return d.page_source

    def close(self):
        try:
            if self._driver:
                self._driver.quit()
        except Exception:
            pass

BROWSER = LazyBrowser()

# ===================== パース =====================
def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    if isinstance(df.columns, pd.MultiIndex):
        cols = []
        for t in df.columns:
            parts = [str(x).strip() for x in t if str(x).strip()]
            s = " ".join(parts) if parts else ""
            toks = s.split()
            if len(toks) >= 2 and len(set(toks)) == 1:
                s = toks[0]
            cols.append(s)
        df.columns = cols
    else:
        df.columns = [str(c).strip() for c in df.columns]

    seen, uniq = {}, []
    for c in df.columns:
        if c in seen:
            seen[c] += 1
            uniq.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            uniq.append(c)
    df.columns = uniq

    if len(df) > 0:
        row0 = [str(v) for v in df.iloc[0].tolist()]
        row0 = [re.sub(r"\s+", "", v) for v in row0]
        col0 = [str(c).replace(" ", "") for c in df.columns]
        match_cnt = sum(1 for x in row0 if x and any(x in c for c in col0 if c))
        if match_cnt >= max(2, len(col0)//2):
            df = df.iloc[1:].reset_index(drop=True)
    return df

REQUIRED_COLS = {"馬番", "人気順", "オッズ", "馬名", "騎手名", "斤量", "性齢"}
COL_PATTERNS = {
    "馬番":   re.compile(r"(馬\s*番|枠\s*番|馬番|枠番|\b馬\s*#?)", re.I),
    "人気順": re.compile(r"(人気|単勝人気)", re.I),
    "オッズ": re.compile(r"(オッズ|単勝)", re.I),
    "馬名":   re.compile(r"(馬\s*名|馬名|名前)", re.I),
    "騎手名": re.compile(r"(騎手|騎手名|ジョッキー)", re.I),
    "斤量":   re.compile(r"(斤量|負担重量|負担重|重量)", re.I),
    "性齢":   re.compile(r"(性\s*齢|性齢|性別?\s*年齢|年齢\s*[／/]\s*性別?)", re.I),
}

def _extract_table(html: str, require_odds: bool = True) -> pd.DataFrame | None:
    """出馬表テーブルを抽出する。

    require_odds=False のときはオッズ・人気順が無くても受理する
    （静的HTMLではJS未実行でオッズが空のため。後段でオッズAPIから補完する）。
    """
    required = REQUIRED_COLS if require_odds else REQUIRED_COLS - {"オッズ", "人気順"}

    # pandas はファイルライクの方が速い
    bio = StringIO(html)
    try:
        tables = pd.read_html(bio, flavor="lxml")
    except Exception:
        bio.seek(0)
        try:
            tables = pd.read_html(bio)
        except Exception:
            return None

    def pick(df: pd.DataFrame) -> pd.DataFrame | None:
        df = _normalize_columns(df)
        cols = [str(c) for c in df.columns]
        mapping = {}
        for want, pat in COL_PATTERNS.items():
            hit = next((c for c in cols if pat.search(c)), None)
            if hit:
                mapping[hit] = want

        # 性と年齢が別カラムの表に対するフォールバック
        if "性齢" not in mapping.values():
            sex_col = next((c for c in cols if re.fullmatch(r"(性|性別)", c, re.I)), None)
            age_col = next((c for c in cols if re.fullmatch(r"(年齢|年令|age)", c, re.I)), None)
            if sex_col and age_col:
                # 一時列を作って性齢として扱う（例: 牡 + 3 → 牡3）
                df["_tmp_性齢"] = (
                    df[sex_col].astype(str).str.extract(r"(牡|牝|セ|騸|騙)", expand=False).fillna("")
                    + df[age_col].astype(str).str.extract(r"(\d+)", expand=False).fillna("")
                )
                mapping["_tmp_性齢"] = "性齢"

        # 足りない時だけ補完（人気/オッズ/騎手/斤量の推定）
        if len(set(mapping.values())) < len(REQUIRED_COLS):
            for c in cols:
                if re.search(r"(印|予想印)", c) and "人気順" not in mapping.values():
                    mapping[c] = "人気順"
                if re.search(r"(単勝|勝率|オッズ)", c, re.I) and "オッズ" not in mapping.values():
                    mapping[c] = "オッズ"
                if re.search(r"(騎手|ジョッキー)", c) and "騎手名" not in mapping.values():
                    mapping[c] = "騎手名"
                if re.search(r"(斤量|負担重量|負担重|重量)", c) and "斤量" not in mapping.values():
                    mapping[c] = "斤量"

        # すべて揃ったら正規化して返す
        if required.issubset(set(mapping.values())):
            out = df[list(mapping.keys())].rename(columns=mapping).copy()
            # オッズ・人気順が無い場合は空列を用意（後段でAPIから補完）
            for c in ("人気順", "オッズ"):
                if c not in out.columns:
                    out[c] = float("nan")

            # ベクトル化正規化
            out["人気順"] = pd.to_numeric(
                out["人気順"].astype(str).str.extract(r"(\d+)", expand=False),
                errors="coerce"
            )
            out["オッズ"] = pd.to_numeric(
                (out["オッズ"].astype(str)
                    .str.replace("倍", "", regex=False)
                    .str.replace(",", "", regex=False)),
                errors="coerce"
            )
            out["馬番"] = (out["馬番"].astype(str)
                .str.extract(r"(\d+)", expand=False).astype("Int64"))
            out["騎手名"] = (out["騎手名"].astype(str)
                .str.replace(r"\s+", " ", regex=True).str.strip())
            out["斤量"] = pd.to_numeric(
                out["斤量"].astype(str).str.extract(r"(\d+(?:\.\d+)?)", expand=False),
                errors="coerce"
            )
            out["馬名"] = out["馬名"].astype(str).str.replace(r"\s+", " ", regex=True).str.strip()

            # 性齢の正規化
            out["性齢"] = out["性齢"].astype(str).str.replace(r"\s+", "", regex=True)

            # 見やすい並びにして返す
            order = [c for c in ["人気順", "馬番", "オッズ", "馬名", "性齢", "斤量", "騎手名"] if c in out.columns]
            return out[order]
        return None

    for tb in tables:
        got = pick(tb)
        if got is not None:
            return got
    return None

PLACE_PATTERN = re.compile(r"(札幌|函館|福島|新潟|東京|中山|中京|京都|阪神|小倉)")

def _extract_race_meta(html: str) -> tuple[str | None, str | None, str | None, str | None, str | None, str | None]:
    """(race_date, name, data01, data02, place, rnum) を返す"""
    soup = BeautifulSoup(html, "lxml")
    name = soup.select_one(".RaceName")
    data01 = soup.select_one(".RaceData01")
    data02 = soup.select_one(".RaceData02")
    rnum  = soup.select_one(".RaceNum")     # 例: 10R

    # テキスト化
    name  = name.get_text(strip=True) if name else None
    data01 = re.sub(r"\s+", " ", data01.get_text(" ", strip=True)) if data01 else None
    data02 = re.sub(r"\s+", " ", data02.get_text(" ", strip=True)) if data02 else None
    rnum  = rnum.get_text(strip=True) if rnum else None
    place: str | None = None

    # rnum 正規化（"第10R" → "10R" など）
    if rnum:
        m = re.search(r"(\d+)\s*R", rnum, flags=re.I)
        if m:
            rnum = f"{int(m.group(1))}R"

    if data02:
        m = PLACE_PATTERN.search(data02)
        if m:
            place = m.group(1)

    # --- race_date 抽出 (yyyymmdd) ---
    race_date: str | None = None

    # 1) 画面上の日本語日付から取得（例: "2025年5月5日"）
    for sel in (".RaceList_Date", ".RaceData01", ".RaceData02"):
        node = soup.select_one(sel)
        if not node:
            continue
        m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", node.get_text(" ", strip=True))
        if m:
            y, mo, d = map(int, m.groups())
            race_date = f"{y:04d}{mo:02d}{d:02d}"
            break

    # 2) ダメなら <script> 内の埋め込み値から拾う（例: "kaisai_date":"20250505"）
    if not race_date:
        for s in soup.find_all("script"):
            st = s.get_text(" ", strip=True)
            m = re.search(r'"kaisai_date"\s*:\s*"(\d{8})"', st) or re.search(r'kaisaiDate\s*[:=]\s*"(\d{8})"', st)
            if m:
                race_date = m.group(1)
                break

    return race_date, name, data01, data02, place, rnum

def _parse_page(html: str) -> tuple[pd.DataFrame, tuple] | None:
    """出馬表テーブルとメタ情報が両方取れたときだけ返す"""
    df = _extract_table(html)
    meta = _extract_race_meta(html)
    _, name, d1, d2, _, _ = meta
    if df is not None and name and d1 and d2:
        return df, meta
    return None

def _fetch_tan_odds(rid: str, timeout_sec: int = 10) -> dict[int, tuple[float, float]] | None:
    """単勝オッズAPIから {馬番: (オッズ, 人気)} を取得する。失敗は None"""
    try:
        r = SESSION.get(ODDS_API_URL.format(rid=rid), timeout=timeout_sec)
        r.raise_for_status()
        tan = r.json()["data"]["odds"]["1"]
        out = {}
        for umaban, vals in tan.items():
            try:
                num = int(umaban)
                odds = float(vals[0])
            except (ValueError, TypeError, IndexError):
                continue
            try:
                ninki = float(vals[2])
            except (ValueError, TypeError, IndexError):
                ninki = float("nan")
            out[num] = (odds, ninki)
        return out or None
    except Exception:
        return None

def _extract_provisional_numbers(html: str) -> dict[str, int]:
    """出馬表の行id（tr_N）から {馬名: 仮馬番} を作る。

    枠順確定前は馬番セルが空だが、行idのNがオッズAPIのキーと一致するため、
    馬名経由でオッズを紐付けられる。
    """
    soup = BeautifulSoup(html, "lxml")
    out = {}
    for tr in soup.select("tr.HorseList"):
        m = re.fullmatch(r"tr_(\d+)", tr.get("id", ""))
        name_el = tr.select_one(".HorseName")
        if m and name_el:
            name = re.sub(r"\s+", " ", name_el.get_text(strip=True)).strip()
            out[name] = int(m.group(1))
    return out

def fetch_static(rid: str, timeout_sec: int = 15) -> tuple[pd.DataFrame, tuple] | None:
    """静的HTML＋オッズAPIでの取得を試みる（並列実行用。失敗は None）"""
    try:
        html = _get_html(SHUTUBA_URL.format(rid=rid), timeout=timeout_sec)
        df = _extract_table(html, require_odds=False)
        meta = _extract_race_meta(html)
        _, name, d1, d2, _, _ = meta
        if df is None or not (name and d1 and d2):
            return None

        # 静的HTMLはオッズがJS描画のため空 → オッズAPIから補完
        if df["オッズ"].isna().all():
            odds_map = _fetch_tan_odds(rid)
            if odds_map:
                nan2 = (float("nan"),) * 2
                if df["馬番"].notna().any():
                    # 枠順確定後: 馬番で紐付け
                    keys = [int(u) if pd.notna(u) else -1 for u in df["馬番"]]
                else:
                    # 枠順確定前: 行id（tr_N）の仮馬番を馬名経由で紐付け
                    prov = _extract_provisional_numbers(html)
                    keys = [prov.get(n, -1) for n in df["馬名"]]
                df["オッズ"] = [odds_map.get(k, nan2)[0] for k in keys]
                df["人気順"] = [odds_map.get(k, nan2)[1] for k in keys]

        # それでもオッズが取れないときは Selenium フォールバックに回す
        if df["オッズ"].isna().all():
            return None
        return df, meta
    except Exception:
        return None

def fetch_rendered(url: str) -> tuple[pd.DataFrame, tuple]:
    """Seleniumでレンダリングして取得（静的取得のフォールバック）"""
    html = BROWSER.get_rendered_html(
        url,
        wait_css=".Shutuba_Table, table.RaceTable01, .RaceTable01",
        hard_timeout=30,
        wait_odds=True
    )
    got = _parse_page(html)
    if got is None:
        raise ValueError("出馬表テーブルが見つかりません。")
    return got

# ===================== WIN5 race_id 抽出（PC→SP フォールバック） =====================
def _extract_ids_from_html(html: str) -> list[str]:
    soup = BeautifulSoup(html, "lxml")
    ids, seen = [], set()
    for a in soup.find_all("a", href=True):
        m = RACE_ID_RE.search(a["href"])
        if not m:
            continue
        rid = m.group(1)
        if rid not in seen:
            seen.add(rid)
            ids.append(rid)
    return ids

def pick_win5_ids(target_url: str | None = None) -> list[str]:
    jst = dt.timezone(dt.timedelta(hours=9))
    today = dt.datetime.now(jst).strftime("%Y%m%d")
    url = target_url or PC_URL

    try:
        ids = _extract_ids_from_html(_get_html(url))
        if len(ids) >= 5:
            return ids[:5]
    except Exception:
        pass

    date = re.search(r"date=(\d{8})", url)
    date = date.group(1) if date else today

    try:
        ids = _extract_ids_from_html(_get_html(SP_URL.format(date=date)))
        return ids[:5] if len(ids) >= 5 else ids
    except Exception:
        return []

# ===================== メタ情報パース =====================
def _parse_race_time(d1) -> str:
    """RaceData01から発走時刻（HH:MM）を抽出する"""
    if not isinstance(d1, str):
        return ""
    m = re.search(r"(\d{1,2}:\d{2})", d1)
    return m.group(1) if m else ""

def _parse_course_label(d1, d2) -> str:
    """〇歳〇メートル（芝・右）形式の文字列を組み立てる"""
    d1 = d1 if isinstance(d1, str) else ""
    d2 = d2 if isinstance(d2, str) else ""

    # 距離（例: 1,800m / 3000m）
    m = re.search(r"([\d,]+)\s*m", d1)
    dist = f"{m.group(1)}メートル" if m else ""

    # 芝/ダート と 右/左/直線（netkeiba表記は「ダ1700m (右)」「芝1800m (右 A)」など）
    m_surf = re.search(r"(障害|障|ダート|ダ|芝)", d1)
    m_dir  = re.search(r"[(（]\s*(右|左|直線)", d1)
    surf_map = {"芝": "芝", "ダ": "ダート", "ダート": "ダート", "障": "障害", "障害": "障害"}
    surf = surf_map.get(m_surf.group(1), "") if m_surf else ""
    drct = m_dir.group(1) if m_dir else ""
    if surf and drct:
        course = f"（{surf}・{drct}）"
    elif surf:
        course = f"（{surf}）"
    else:
        course = ""

    # 年齢条件（例: 4歳以上 / 3歳）をd2から抽出
    m_age = re.search(r"(\d+歳[以上未満]*)", d2)
    age = m_age.group(1) if m_age else ""

    # クラス条件（例: 3勝クラス / オープン / G1 など）
    m_cls = re.search(r"(\d+勝クラス|オープン|G[IⅠ1iIVX]{1,3}|ハンデ|混合|牝馬限定)", d2)
    cls = m_cls.group(1) if m_cls else ""

    return f"{age}{cls}{dist}{course}".strip()

# ===================== テンプレートへの書き込み =====================
def write_race_to_odds_sheet(ws, win_idx: int, df: pd.DataFrame, race_title: str, race_time: str, course_label: str):
    """オッズデータ入力シートの指定WIN区画にデータを書き込む"""
    sec = WIN_SECTION_COLS[win_idx]

    ws.cell(row=TIME_ROW,      column=sec).value = race_time
    ws.cell(row=RACE_NAME_ROW, column=sec).value = race_title
    ws.cell(row=COURSE_ROW,    column=sec).value = course_label

    # 既存データをクリア（数式・結合セルはスキップ）
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        for off in DATA_COL_OFFSETS.values():
            cell = ws.cell(row=row, column=sec + off)
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = None

    # データを書き込む
    for r_off, row_vals in enumerate(df.itertuples(index=False)):
        r = DATA_START_ROW + r_off
        if r > DATA_END_ROW:
            break
        for col_name, value in zip(df.columns, row_vals):
            if col_name not in DATA_COL_OFFSETS:
                continue
            # NaN / pd.NA（枠順確定前の馬番など）は空セルにする
            if value is not None and not isinstance(value, str) and pd.isna(value):
                value = None
            cell = ws.cell(row=r, column=sec + DATA_COL_OFFSETS[col_name])
            if isinstance(cell, MergedCell):
                continue
            if isinstance(cell.value, str) and cell.value.startswith("="):
                continue
            cell.value = value

def get_output_dir() -> Path:
    try:
        base = Path(__file__).resolve().parent
    except NameError:
        base = Path.cwd()
    out = base / "output"
    out.mkdir(parents=True, exist_ok=True)
    return out

# ===================== メイン =====================
def main():
    t_start = time.time()
    url_arg = sys.argv[1] if len(sys.argv) >= 2 else None
    race_ids = pick_win5_ids(url_arg)
    if not race_ids:
        print("対象の race_id を取得できませんでした。")
        sys.exit(2)

    if not TEMPLATE_XLSX.exists():
        print(f"テンプレートが見つかりません: {TEMPLATE_XLSX}")
        sys.exit(3)

    nowstamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    outdir = get_output_dir()
    out_xlsx = outdir / f"Win5出馬表_{nowstamp}.xlsx"
    print(f"出力開始: {out_xlsx}")

    wb = load_workbook(TEMPLATE_XLSX)
    ws_odds = wb["オッズデータ入力"]

    # 静的HTML＋オッズAPIを並列取得（失敗したレースだけ後段でSeleniumフォールバック）
    race_ids = race_ids[:len(WIN_SECTION_COLS)]
    print(f"{len(race_ids)}レースを並列取得中…")
    with ThreadPoolExecutor(max_workers=len(race_ids)) as ex:
        results = list(ex.map(fetch_static, race_ids))

    errors = []
    written = 0

    for idx_r, (rid, got) in enumerate(zip(race_ids, results)):
        try:
            if got is None:
                print(f"第{idx_r+1}レース: 静的取得失敗 → Seleniumで再取得中…")
                got = fetch_rendered(SHUTUBA_URL.format(rid=rid))

            df, meta = got
            _, name, d1, d2, place, rnum = meta

            race_title   = f"{place}{rnum}_{name}" if place and rnum else name
            race_time    = _parse_race_time(d1)
            course_label = _parse_course_label(d1, d2)
            keys = [c for c in ["人気順", "馬番"] if c in df.columns]
            if keys:
                df = df.sort_values(keys, na_position="last", ignore_index=True, kind="mergesort")

            write_race_to_odds_sheet(ws_odds, idx_r, df, race_title, race_time, course_label)
            print(f"第{idx_r+1}レース [{race_title}] 書き込み完了（{len(df)}頭）")
            written += 1
        except Exception as e:
            msg = f"{rid}: {type(e).__name__}: {e}"
            print("[SKIP]", msg)
            errors.append(msg)

    wb.save(out_xlsx)
    BROWSER.close()
    print(f"出力完了: {out_xlsx}（{written}/{len(race_ids)}レース、{time.time() - t_start:.1f}秒）")
    if errors:
        print("スキップしたレース:")
        for msg in errors:
            print(" -", msg)

if __name__ == "__main__":
    main()
